#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
智能CAD报价生成器 - 美观版本
"""

import streamlit as st
import openpyxl
import ezdxf
from copy import copy
import tempfile
import os

# ========== 辅助函数 ==========
def find_matching_project(layer_name):
    """根据图层名称找到匹配的工程项目"""
    layer_mapping = {
        '活动家具': '家具安装',
        '房门': '套装门安装',
        '新建墙体': '石膏板隔墙（含造型）',
        '拆墙': '拆除墙体',
        '地面': '抛釉砖铺设',
        '地砖': '抛釉砖铺设',
        '墙面': '腻子批刮',
        '吊顶': '石膏板平面顶',
        '灯具': '筒灯/射灯开孔安装',
        '梁': '梁面处理',
        '门槛石': '门槛石安装',
        '窗帘箱': '暗藏窗帘箱（直线型）',
        '空调框架': '空调出风、回风口框架制作',
        '挡水条': '挡水条安装',
        '卫生间': '防水涂料施工',
        '厨房': '橱柜安装',
        '阳台': '地砖铺设',
    }

    if layer_name in layer_mapping:
        return layer_mapping[layer_name]

    for key, project in layer_mapping.items():
        if key in layer_name or layer_name in key:
            return project

    return None

# ========== 页面配置 ==========
st.set_page_config(
    page_title="智能CAD报价生成器",
    page_icon="🏗️",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ========== 自定义CSS ==========
st.markdown("""
<style>
    /* 隐藏默认样式 */
    .stApp .deployButton { display: none; }

    /* 侧边栏样式 */
    .css-1d391kg { background-color: #f8f9fa; }

    /* 文件上传器样式 */
    .stFileUploader { border: 3px dashed #667eea; border-radius: 15px; padding: 30px; }

    /* 按钮样式 */
    .stButton > button { background-color: #667eea; color: white; border: none; padding: 15px 40px; border-radius: 30px; font-size: 1.2rem; font-weight: bold; }
    .stButton > button:hover { background-color: #764ba2; }
</style>
""", unsafe_allow_html=True)

# ========== 主标题 ==========
st.markdown("""
<div style='background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); padding: 50px 30px; border-radius: 20px; margin-bottom: 30px; text-align: center;'>
    <h1 style='color: white; font-size: 3rem; margin: 0;'>🏗️ 智能CAD报价生成器</h1>
    <p style='color: rgba(255,255,255,0.9); font-size: 1.3rem; margin: 15px 0 0 0;'>三步完成，自动生成专业报价单</p>
</div>
""", unsafe_allow_html=True)

# ========== 侧边栏 ==========
with st.sidebar:
    st.markdown("""
    <div style='background: white; padding: 25px; border-radius: 15px; box-shadow: 0 4px 15px rgba(0,0,0,0.1); margin-bottom: 20px;'>
        <h2 style='color: #667eea; font-size: 1.5rem; margin-bottom: 20px;'>📖 使用说明</h2>
        <ol style='color: #666; line-height: 2; padding-left: 20px;'>
            <li>上传价格模板Excel</li>
            <li>上传样式模板Excel</li>
            <li>上传DXF图纸文件</li>
            <li>点击生成报价单</li>
            <li>下载Excel文件</li>
        </ol>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # 状态面板
    st.markdown("""
    <div style='background: white; padding: 25px; border-radius: 15px; box-shadow: 0 4px 15px rgba(0,0,0,0.1);'>
        <h2 style='color: #667eea; font-size: 1.5rem; margin-bottom: 20px;'>📊 上传状态</h2>
    """, unsafe_allow_html=True)

    if 'price_template' in st.session_state:
        st.success("✅ 价格模板已上传")
        st.caption(f"📄 {st.session_state.price_template.name}")
    else:
        st.warning("⏳ 等待上传价格模板")
        st.caption("请上传包含单价的Excel文件")

    st.markdown("<br>", unsafe_allow_html=True)

    if 'style_template' in st.session_state:
        st.success("✅ 样式模板已上传")
        st.caption(f"📄 {st.session_state.style_template.name}")
    else:
        st.warning("⏳ 等待上传样式模板")
        st.caption("请上传包含格式的Excel文件")

    st.markdown("<br>", unsafe_allow_html=True)

    if 'dxf_file' in st.session_state:
        st.success("✅ DXF图纸已上传")
        st.caption(f"📐 {st.session_state.dxf_file.name}")
    else:
        st.warning("⏳ 等待上传DXF图纸")
        st.caption("请上传CAD绘制的DXF文件")

    st.markdown("</div>", unsafe_allow_html=True)

# ========== 主内容区 ==========
st.markdown("""
<div style='background: white; padding: 30px; border-radius: 15px; box-shadow: 0 4px 15px rgba(0,0,0,0.1);'>
""", unsafe_allow_html=True)

# 步骤1和2：上传模板
st.markdown("<h2 style='color: #667eea; margin-bottom: 20px;'>📁 第一步：上传模板</h2>", unsafe_allow_html=True)

col1, col2 = st.columns(2)

with col1:
    st.markdown("""
    <div style='background: #f8f9fa; padding: 20px; border-radius: 10px; border-left: 4px solid #667eea; margin-bottom: 15px;'>
        <h3 style='color: #667eea; margin: 0 0 10px 0;'>📋 价格模板Excel</h3>
        <p style='color: #666; margin: 0;'>包含单价信息的Excel文件</p>
    </div>
    """, unsafe_allow_html=True)

    price_template = st.file_uploader(
        "👆 点击或拖拽上传价格模板",
        type=['xlsx', 'xls'],
        key='price_template_upload',
        label_visibility="collapsed"
    )

    if price_template:
        st.session_state.price_template = price_template
        st.success(f"✅ 已上传: **{price_template.name}**")
        st.caption(f"📊 文件大小: {price_template.size / 1024:.1f} KB")

with col2:
    st.markdown("""
    <div style='background: #f8f9fa; padding: 20px; border-radius: 10px; border-left: 4px solid #667eea; margin-bottom: 15px;'>
        <h3 style='color: #667eea; margin: 0 0 10px 0;'>🎨 样式模板Excel</h3>
        <p style='color: #666; margin: 0;'>包含格式的报价单Excel文件</p>
    </div>
    """, unsafe_allow_html=True)

    style_template = st.file_uploader(
        "👆 点击或拖拽上传样式模板",
        type=['xlsx', 'xls'],
        key='style_template_upload',
        label_visibility="collapsed"
    )

    if style_template:
        st.session_state.style_template = style_template
        st.success(f"✅ 已上传: **{style_template.name}**")
        st.caption(f"📊 文件大小: {style_template.size / 1024:.1f} KB")

st.markdown("<hr style='border: none; border-top: 2px solid #e9ecef; margin: 40px 0;'>", unsafe_allow_html=True)

# 步骤3：上传DXF
st.markdown("<h2 style='color: #667eea; margin-bottom: 20px;'>📐 第二步：上传DXF图纸</h2>", unsafe_allow_html=True)

st.markdown("""
<div style='background: #f8f9fa; padding: 20px; border-radius: 10px; border-left: 4px solid #667eea; margin-bottom: 15px;'>
    <h3 style='color: #667eea; margin: 0 0 10px 0;'>🖼️ DXF图纸文件</h3>
    <p style='color: #666; margin: 0;'>CAD绘制的DXF图纸文件</p>
</div>
""", unsafe_allow_html=True)

dxf_file = st.file_uploader(
    "👆 点击或拖拽上传DXF图纸",
    type=['dxf'],
    key='dxf_file_upload',
    label_visibility="collapsed"
)

if dxf_file:
    st.session_state.dxf_file = dxf_file
    st.success(f"✅ 已上传: **{dxf_file.name}**")
    st.caption(f"📊 文件大小: {dxf_file.size / 1024:.1f} KB")

st.markdown("<hr style='border: none; border-top: 2px solid #e9ecef; margin: 40px 0;'>", unsafe_allow_html=True)

# 步骤4：项目信息
st.markdown("<h2 style='color: #667eea; margin-bottom: 20px;'>📝 第三步：填写项目信息（可选）</h2>", unsafe_allow_html=True)

col3, col4 = st.columns(2)

with col3:
    project_name = st.text_input(
        "🏠 项目名称",
        value="住宅装修",
        placeholder="例如：甘长名苑2-1902"
    )

with col4:
    client_address = st.text_input(
        "📍 工程地址",
        value="",
        placeholder="例如：杭州市西湖区xxx路xxx号"
    )

st.markdown("<hr style='border: none; border-top: 2px solid #e9ecef; margin: 40px 0;'>", unsafe_allow_html=True)

# 步骤5：生成按钮
st.markdown("<h2 style='color: #667eea; margin-bottom: 20px;'>🚀 第四步：生成报价单</h2>", unsafe_allow_html=True)

all_uploaded = ('price_template' in st.session_state and
                'style_template' in st.session_state and
                'dxf_file' in st.session_state)

if not all_uploaded:
    st.info("⏳ 请先上传所有必需的文件（价格模板 + 样式模板 + DXF图纸）")

st.markdown("<div style='text-align: center; margin: 30px 0;'>", unsafe_allow_html=True)

generate_btn = st.button(
    "🚀 开始生成报价单",
    disabled=not all_uploaded,
    use_container_width=True
)

st.markdown("</div>", unsafe_allow_html=True)

st.markdown("</div>", unsafe_allow_html=True)

# ========== 生成逻辑 ==========
if generate_btn:
    # 显示进度
    with st.spinner("🔄 正在生成报价单，请稍候..."):
        progress_bar = st.progress(0)
        status_text = st.empty()

        try:
            # 1. 保存文件
            status_text.markdown("📁 正在保存上传的文件...")
            temp_dir = tempfile.mkdtemp()

            price_template_path = os.path.join(temp_dir, "price_template.xlsx")
            style_template_path = os.path.join(temp_dir, "style_template.xlsx")
            dxf_path = os.path.join(temp_dir, "drawing.dxf")

            with open(price_template_path, 'wb') as f:
                f.write(st.session_state.price_template.read())

            with open(style_template_path, 'wb') as f:
                f.write(st.session_state.style_template.read())

            with open(dxf_path, 'wb') as f:
                f.write(st.session_state.dxf_file.read())

            progress_bar.progress(20)

            # 2. 解析DXF
            status_text.markdown("📐 正在解析DXF图纸...")
            doc = ezdxf.readfile(dxf_path)
            msp = doc.modelspace()

            layer_counts = {}
            for entity in msp:
                layer = entity.dxf.layer
                layer_counts[layer] = layer_counts.get(layer, 0) + 1

            st.info(f"📊 发现 **{len(layer_counts)}** 个图层")
            progress_bar.progress(40)

            # 3. 加载价格
            status_text.markdown("💰 正在加载价格模板...")
            wb_price = openpyxl.load_workbook(price_template_path)
            ws_price = wb_price.active

            prices = {}
            # 从第9行开始读取数据（前8行是标题）
            for i in range(9, ws_price.max_row + 1):
                code = ws_price.cell(i, 2).value  # B列：定额编号
                project = ws_price.cell(i, 4).value  # D列：工程项目
                unit = ws_price.cell(i, 6).value  # F列：单位

                # 读取主材、辅料、机械、人工、损耗（价格模板可能有这些列）
                # 报价书格式：G(7)-主材, H(8)-辅料, I(9)-机械, J(10)-人工, K(11)-损耗
                material = ws_price.cell(i, 7).value  # G列
                auxiliary = ws_price.cell(i, 8).value  # H列
                mechanical = ws_price.cell(i, 9).value  # I列
                labor = ws_price.cell(i, 10).value  # J列
                waste = ws_price.cell(i, 11).value  # K列

                if project and isinstance(project, str) and code:
                    try:
                        # 计算单价 = 主材+辅料+机械+人工+损耗
                        price_num = 0
                        for val in [material, auxiliary, mechanical, labor, waste]:
                            if val is not None:
                                price_num += float(val)

                        prices[project] = {
                            'code': code,
                            'project': project,
                            'unit': unit,
                            'material': material if material is not None else 0,
                            'auxiliary': auxiliary if auxiliary is not None else 0,
                            'mechanical': mechanical if mechanical is not None else 0,
                            'labor': labor if labor is not None else 0,
                            'waste': waste if waste is not None else 0,
                            'price': price_num  # 单价 = 各项之和
                        }
                    except (ValueError, TypeError):
                        pass

            wb_price.close()
            st.success(f"💰 已加载 **{len(prices)}** 个工程项目")
            progress_bar.progress(60)

            # 4. 生成Excel
            status_text.markdown("📋 正在生成报价单...")
            template_wb = openpyxl.load_workbook(style_template_path)

            new_wb = openpyxl.Workbook()
            if 'Sheet' in new_wb.sheetnames:
                del new_wb['Sheet']

            for sheet_name in template_wb.sheetnames:
                src_ws = template_wb[sheet_name]
                dst_ws = new_wb.create_sheet(sheet_name)

                for i in range(1, src_ws.max_row + 1):
                    dst_ws.row_dimensions[i].height = src_ws.row_dimensions[i].height
                    for j in range(1, src_ws.max_column + 1):
                        src_cell = src_ws.cell(i, j)
                        dst_cell = dst_ws.cell(i, j)
                        dst_cell.value = src_cell.value
                        if src_cell.has_style:
                            dst_cell.font = copy(src_cell.font)
                            dst_cell.border = copy(src_cell.border)
                            dst_cell.fill = copy(src_cell.fill)
                            dst_cell.number_format = src_cell.number_format
                            dst_cell.alignment = copy(src_cell.alignment)

                for j in range(1, src_ws.max_column + 1):
                    col_letter = openpyxl.utils.get_column_letter(j)
                    dst_ws.column_dimensions[col_letter].width = src_ws.column_dimensions[col_letter].width

                for merge_range in src_ws.merged_cells.ranges:
                    try:
                        dst_ws.merge_cells(merge_range.coord)
                    except:
                        pass

            progress_bar.progress(80)

            # 5. 填充数据
            if '预算' in new_wb.sheetnames:
                sheet_budget = new_wb['预算']
                if client_address:
                    sheet_budget.cell(4, 1).value = f"工程地址：{client_address}"

                total_price = 0
                data_row = 8
                section_end = 1000  # 放大限制

                for layer_name, count in sorted(layer_counts.items(), key=lambda x: -x[1]):
                    if data_row >= section_end:
                        break

                    project_info = find_matching_project(layer_name)

                    if project_info and project_info in prices:
                        price_info = prices[project_info]

                        # 按照报价书格式填充
                        # A列(1): 编号
                        sheet_budget.cell(data_row, 1).value = price_info['code']

                        # B列(2): 工程项目
                        sheet_budget.cell(data_row, 2).value = price_info['project']

                        # C列(3): 单位
                        sheet_budget.cell(data_row, 3).value = price_info['unit']

                        # D列(4): 数量（从DXF图层获取）
                        sheet_budget.cell(data_row, 4).value = count

                        # E列(5): 单价（=G+H+I+J+K，即主材+辅料+机械+人工+损耗）
                        sheet_budget.cell(data_row, 5).value = price_info['price']

                        # F列(6): 金额（=单价*数量）
                        item_total = price_info['price'] * count
                        sheet_budget.cell(data_row, 6).value = item_total

                        # G列(7): 主材
                        sheet_budget.cell(data_row, 7).value = price_info['material']

                        # H列(8): 辅料
                        sheet_budget.cell(data_row, 8).value = price_info['auxiliary']

                        # I列(9): 机械
                        sheet_budget.cell(data_row, 9).value = price_info['mechanical']

                        # J列(10): 人工
                        sheet_budget.cell(data_row, 10).value = price_info['labor']

                        # K列(11): 损耗
                        sheet_budget.cell(data_row, 11).value = price_info['waste']

                        total_price += item_total
                        data_row += 1

            output_path = os.path.join(temp_dir, "报价单.xlsx")
            new_wb.save(output_path)
            template_wb.close()
            new_wb.close()

            progress_bar.progress(100)

            # 成功提示
            st.markdown("<br>", unsafe_allow_html=True)
            st.markdown("""
            <div style='background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); padding: 30px; border-radius: 15px; text-align: center;'>
                <h2 style='color: white; margin: 0;'>✅ 报价单生成成功！</h2>
            </div>
            """, unsafe_allow_html=True)

            st.markdown(f"""
            <div style='background: #d4edda; padding: 25px; border-radius: 10px; border: 2px solid #c3e6cb; text-align: center; margin: 20px 0;'>
                <h1 style='color: #155724; margin: 0;'>💰 总价: ￥{total_price:,.2f}</h1>
            </div>
            """, unsafe_allow_html=True)

            # 下载按钮
            with open(output_path, 'rb') as f:
                st.download_button(
                    label="💾 下载报价单",
                    data=f,
                    file_name="报价单.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

            # 图层统计
            st.markdown("<br>", unsafe_allow_html=True)
            st.markdown("<h2 style='color: #667eea; margin-bottom: 20px;'>📊 图层统计</h2>", unsafe_allow_html=True)

            layer_data = []
            for layer, count in sorted(layer_counts.items(), key=lambda x: -x[1]):
                project_info = find_matching_project(layer)
                if project_info and project_info in prices:
                    price = prices[project_info]['price']
                    total = price * count
                    layer_data.append({
                        '图层名称': layer,
                        '工程项目': project_info,
                        '数量': count,
                        '单价': f'￥{price:,.2f}',
                        '总价': f'￥{total:,.2f}'
                    })

            if layer_data:
                st.dataframe(layer_data, use_container_width=True, hide_index=True)

        except Exception as e:
            st.markdown(f"""
            <div style='background: #f8d7da; padding: 20px; border-radius: 10px; border: 2px solid #f5c6cb;'>
                <h3 style='color: #721c24; margin: 0 0 10px 0;'>❌ 生成失败</h3>
                <p style='color: #721c24; margin: 0;'>{str(e)}</p>
            </div>
            """, unsafe_allow_html=True)
            import traceback
            st.error(traceback.format_exc())


# ========== 底部 ==========
st.markdown("<br>", unsafe_allow_html=True)
st.markdown("""
<div style='text-align: center; color: #999; padding: 20px;'>
    <p>💡 生成的Excel文件格式与样式模板完全一致</p>
    <p>📧 如有问题，请联系开发者</p>
</div>
""", unsafe_allow_html=True)
